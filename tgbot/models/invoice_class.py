# -*- coding: utf-8 -*-
# Не оптимизировать импорты и не менять их порядок

import os, django

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "dj_ac.settings")
os.environ.update({'DJANGO_ALLOW_ASYNC_UNSAFE': "true"})
django.setup()

import logging
from asgiref.sync import sync_to_async
from invoice.models import Invoice
logger = logging.getLogger(__name__)

from num2words import num2words
import math
from datetime import *
from tgbot.data import config


@sync_to_async()
class Invoice:
    print("class Invoice:")
    """Класс Invoice используется для создания Счетов на оплату
        Основное применение - формирование Счета и экспорт данных в формат xlsx
        Note:
            Все атрибуты заданы по-умолчанию

        Attributes
        ----------
        quantity : int
            количество товара или услуги
        name_of_services : str
            название услуги
        customer : str
            наименование покупателя (организации)
        executor: str
            Название и реквизиты исполнителя
        units: str
            единицы измерения (шт, м.кв.)
        price: int
            цена одной единицы услуги
        nds: str
            прописывается поле с НДС
        summa: int
            считается автоматически
        names: int
            общее кол-во наименований
        director: str
            директор

        Methods
        -------
        exel_generator()
            Создает эксель файл со Счетом
        summa_propisyu ()
            Приводит число к принятому в документах текстовому отображению
        names_string_result()
            Выводит строку 'Счет на оплату №...от....'
        date_words()
            Формирует дату в стандартном для счета представлении
        number()
            Формирует номер счета, прибавляя единицу к данным из счетчика
        """

    def __init__(self,
                 customer='ИП Телелинский Дмитрий Владимирович  ,'
                          ' ИНН  - , КПП   -, 350063, город Краснодар,'
                          ' ул.Кубанская Набережная, д.37, кв.140',
                 executor='ООО "Неолюкс", ИНН 6658215373, КПП 231001001, 350010,'
                          ' г. Краснодар, ул. Зиповская, д. 5, Телефон: +7 (978) 074-08-54 ',
                 name_of_services='Дизайн проект по Договору', units='кв.м.', quantity: int = 100, price: int = 400,
                 nds='Без НДС', names=1,
                 director='Кандалов В.А.',
                 bookkeeper='Кандалов В.А.',
                 number: int = None,
                 number_akt: str = "Акт № 03 от ",
                 name_of_invoice: str = "Счёт № 03 от",):
        self.customer = customer  # Заказчик (достается из словаря с заказчиками или из БД)
        self.executor = executor  # Исполнитель
        self.name_of_services = name_of_services  # Наименование услуг
        self.units = units  # Единицы измерения
        self.quantity = quantity  # Количество
        self.price = price  # Цена
        self.summa = '{0:.2f}'.format(int(self.price) * int(self.quantity))  # Сумма
        self.prefix = int(self.price) * int(self.quantity)
        self.data = datetime.now()  # Дата счета
        self.nds = nds
        self.names = names
        self.director = director
        self.bookkeeper = bookkeeper
        self.number = number
        self.number_akt = number_akt
        self.name_of_invoice = name_of_invoice

    def __repr__(self):  # Магический дандр метод форматирования вывода print(p): Дело(Название=1, Сделано?=1)
        return f'Документ: {self.number} {self.date_words()}\n' \
               f'   Заказчик: {self.customer}\n' \
               f'   Исполнитель: {self.executor}\n' \
               f'   Сумма: {self.summa} рублей \n' \
               f'   Основание: {self.name_of_services}\n'

    def exel_generator(self):
        print("def exel_generator(self):")
        from openpyxl import load_workbook
        wb = load_workbook(config.EXEL_PATH2 + 'example2.xlsx')
        wb_filename = f'{config.EXEL_PATH2}invoice_{self.prefix}_{self.number_akt}.xlsx'
        ws = wb["Заготовка счета"]
        ws['U22'] = self.quantity  # Количество
        ws['D22'] = self.name_of_services  # Товар
        ws['AB22'] = self.price  # Цена
        ws['Y22'] = self.units  # Единицы измерения
        ws['F19'] = self.customer  # Покупатель
        ws['AH22'] = self.summa  # Сумма
        ws['AH24'] = self.summa  # Итого
        ws['AH26'] = self.summa  # Всего к оплате
        ws['B27'] = self.names_string_result()  # Всего наименований
        ws['B28'] = self.summa_propisyu()  # Сумма прописью
        ws['B13'] = f'Счёт № {self.number_akt} от {self.date_words()}'  # Номер счета
        wb.save(filename=wb_filename)
        return wb_filename

    def exel_generator_akt(self):
        print("def exel_generator_akt(self):")
        from openpyxl import load_workbook
        wb = load_workbook(config.EXEL_PATH2 + 'example2.xlsx')
        wb_filename_akt = f'{config.EXEL_PATH2}akt_{self.prefix}_{self.number_akt}.xlsx'
        ws = wb["Заготовка акта"]
        ws['U11'] = self.quantity  # Количество
        ws['D11'] = self.name_of_services  # Товар
        ws['Z11'] = self.price  # Цена
        ws['X11'] = self.units  # Единицы измерения
        ws['F7'] = self.customer  # Покупатель
        ws['AD11'] = self.summa  # Сумма
        ws['AD13'] = self.summa  # Итого
        ws['B16'] = self.names_string_result()  # Всего наименований
        ws['B17'] = self.summa_propisyu()  # Сумма прописью
        ws['B3'] = f'Акт № {self.number_akt} от {self.date_words()}'  # Номер Акта

        wb.save(filename=wb_filename_akt)
        return wb_filename_akt

    def summa_propisyu(self):
        kops = int(self.price) * float(self.quantity)
        kopeyki = math.modf(kops)[0]
        rubli = math.modf(float(self.summa))[1]
        rubliround = int(rubli)
        word3 = num2words(str(rubliround), lang='ru')
        wrd31 = (word3.capitalize())
        newkopeyki = str(kopeyki)
        newkopeyki2 = newkopeyki.replace('.', "")
        return f'{str(wrd31)} рублей {newkopeyki2} копеек. НДС 0%.'

    def names_string_result(self):
        return f'Всего наименований {self.names}, на сумму {self.summa} руб.'

    @staticmethod
    def number():
        return f'Счет на оплату № {1} от '

    @staticmethod
    def number_akt():
        return f'Акт № {1} от '

    @staticmethod
    def date_words():
        a = ['Января', 'Февраля', 'Марта', 'Апреля', 'Мая', 'Июня', 'Июля', 'Августа', 'Сентября', 'Октября', 'Ноября',
             'Декабря']
        a1 = ['01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12', '13', '14', '15', '16', '17',
              '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31']
        d = datetime.today()
        g = d.day
        c = d.month
        h = d.year

        return f'{a1[g - 1]} {a[c - 1]} {str(h)}'


def invoce_maker():
    new_invoice = Invoice(price=500, quantity=989)
    new_invoice.exel_generator()
    new_invoice2 = Invoice(price=600, quantity=700)
    new_invoice2.exel_generator_akt()


if __name__ == '__main__':
    invoce_maker()
